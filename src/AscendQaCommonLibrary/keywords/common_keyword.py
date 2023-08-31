from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from robot.libraries.BuiltIn import BuiltIn
from robot.libraries.Collections import Collections
from selenium.webdriver.remote.webelement import WebElement
from SeleniumLibrary.base import LibraryComponent, keyword
from SeleniumLibrary import ScreenshotKeywords
from SeleniumLibrary.utils.path_formatter import _format_path

import random, string
import os
import platform
import os.path
import sys
from PyPDF2 import PdfReader
from AscendQaCommonLibrary.utils import ImageUtils
from AscendQaCommonLibrary.utils import GeneralUtils
from AscendQaCommonLibrary.utils import DateUtils

from datetime import datetime, timezone, timedelta,date
import datetime
from openpyxl import Workbook , load_workbook
from ImapLibrary2 import ImapLibrary2

# keywords = SeleniumLibrary().get_keyword_names()

##### for log ####
# print({length})
# print(f"{keywords}")
# BuiltIn().log_to_console(rand_letters)

class common_keyword(LibraryComponent):
    @keyword
    def get_driver_instance(self):
        return BuiltIn().get_library_instance('SeleniumLibrary')
    
    # Get thai year from current AD year
    # ``year`` is AD year. ex. 2022 
    @keyword
    def get_thai_year(self,year:int):
        year_thai = year + 543
        BuiltIn().log_to_console(year_thai)
        return year_thai
    
    # Get day of the week in thai
    # ``day``  as Sunday will return `อาทิตย์`
    @keyword
    def get_day_of_week_in_thai(self,day:string):
        day = day.lower()
        BuiltIn().log_to_console(day.lower())
        BuiltIn().log_to_console(day)
        date_thai = ''
        if    day == 'sunday':
            date_thai =  'อาทิตย์'
        elif    day == 'monday':
            date_thai = 'จันทร์'
        elif    day == 'tuesday':
            date_thai =  'อังคาร'
        elif    day == 'wednesday':
            date_thai = 'พุธ'
        elif    day == 'thursday':
            date_thai =  'พฤหัสบดี'
        elif    day == 'friday':
            date_thai = 'ศุกร์'
        elif    day == 'saturday':
            date_thai = 'เสาร์'

        return date_thai
    
    @keyword
    def generate_random_number(self,length:int):
        number = '0123456789'
        rand_num = ''.join(random.choice(number) for i in range(length))
        return rand_num
    
    @keyword
    def generate_random_string(self,length:int):
        letters = string.ascii_lowercase
        rand_letters = ''.join(random.choice(letters) for i in range(length))
        return rand_letters
    
    # Random thai mobile number start with 08
    @keyword
    def random_thai_mobile_number(self):
        thai_mobile_start = '08'
        number = '0123456789'
        rand_num = ''.join(random.choice(number) for i in range(8))
        mobile_number = thai_mobile_start + rand_num
        return mobile_number
    
    # For generate random thai id card 13.
    # ...    random 12 number and find number of position 13
    # ...    sum of multiply number with position by start position at 13 ex. sum = (n1 * 13) + (n2 * 12) + (n3 * 11)...
    # ...    sum mod 11
    # ...    So find position 13 = 11 - value of mod (if value 11 use 1 ,but value 10 use 0)
    @keyword
    def random_thai_citizen_id(self):
        random_num = self.generate_random_number(12)
        length = len(random_num)
        position = 13
        sum = 0
        total = 0
        num_13 = 0
        position_13 = 0
        for index in range(length):
            p = position - index
            sum = random_num[index] * p
            total = int(total)+int(sum)
            mod = total%11
            num_13 = 11-mod
            if    'num_13' == '10':
                position_13 =  0
            elif    'num_13' == '11':
                position_13 =  1
            else:
                position_13 =  num_13    

            citizen_id = random_num + str(position_13)
        return citizen_id
    @keyword
    def verify_two_numbers_equal(self,number1:int, number2:int):
        if number1 == number2:
            print(f"{number1} equal {number2}")
        else:
            raise AssertionError(f"{number1} not equal {number2}")

    # Get os platform 
    #     ...     \n either darwin (Mac) or window (Ex. Darwin -> darwin)
    @keyword
    def get_os_platform(self):
        # BuiltIn().log_to_console(sys.platform())
        plat_form = platform.system()
        system = plat_form.lower()
        return    system
    
    # Get normalize path
    # Normalizes the given path.
    #     ...     \n Collapses redundant separators and up-level references.
    #     ...     \n Converts `/` to `\` on Windows.
    #     ...     \n Replaces initial ~ or ~user by that user's home directory.
    @keyword
    def get_normalize_path(self,path):
        normal_path = os.path.normpath(path)
        return    normal_path
    
    # PDF should contain
    # Verify if pdf file contains expected message or not
    @keyword
    def pdf_should_contain(self,pdf_path,message):
        path = self.get_normalize_path(pdf_path)

        # creating a pdf reader object
        reader = PdfReader(path)

        # printing number of pages in pdf file
        # print(len(reader.pages))
        
        # getting a specific page from the pdf file
        page = reader.pages[0]

        # extracting text from page
        txt = page.extract_text()

        BuiltIn().should_contain(txt,message)

        # Image should be visible on screen   
#     [Documentation]     Find image on current screen by comparing screenshot of current screen and expected image
#     ...     \n default threshold is 0.8 meaning 80% of 2 images should match
#     ...     \n Return true/false and xy of the expected image on screen if any
#     ...     \n ``abs_expected_image_path`` is the absolute path of expected image
#     [Arguments]     ${abs_expected_image_path}   ${threshold}=0.8
#     ${current_time}=            BuiltIn.Get time    epoch
#     ${screen_screenshot}=       SeleniumLibrary.Capture Page Screenshot     ${OUTPUT_DIR}${/}screen_screenshot_${current_time}.png
#     ${is_found_image}   ${xy}=  ImageUtils.Image should be visible on screen
#                                 ...     ${abs_expected_image_path}
#                                 ...     ${screen_screenshot}
#                                 ...     ${threshold}
#     [Return]    ${is_found_image}   ${xy}
    @keyword
    def image_should_be_visible_on_screen(self,abs_expected_image_path,threshold=0.8):
        driver = self.get_driver_instance()
        current_time = BuiltIn().get_time('epoch')
        BuiltIn().log_to_console(current_time)
        filename = str('/screen_screenshot_'+ str(current_time) +'.png')
        BuiltIn().log_to_console(filename)
        screen_screenshot = driver.capture_page_screenshot(filename)
        # BuiltIn().log_to_console(screen_screenshot)
        is_found_image,xy= ImageUtils.image_should_be_visible_on_screen(abs_expected_image_path,screen_screenshot,threshold)
        # BuiltIn().log_to_console(is_found_image)
        # BuiltIn().log_to_console(xy)
        return    is_found_image,xy

    # Wait until download is completed
    #Verifies that the directory has one or more folder and it is not a temp file.
    #\n returns path to the file
    #\n `retry`  how many time to loop retry
    #\n `wait_time`  how many second to wait before start another retry
    #\n Current version is able to find new file when folder has more one file
    @keyword
    def wait_until_download_is_completed(self,directory,retry=5,wait_time=2):
        list_of_existing_files = os.listdir(directory)
        total_number_of_existing_files = len(list_of_existing_files)
        expected_number_of_file_after_download = total_number_of_existing_files+1
        new_files_name = ''
        is_not_temp_file = False
        for index in retry:
            list_of_current_files = os.listdir(directory)
            number_of_current_files = len(list_of_current_files)
            is_contains_new_file = BuiltIn().run_keyword_and_return_status(BuiltIn().should_be_equal(number_of_current_files,expected_number_of_file_after_download,'Should be' + expected_number_of_file_after_download + 'file in the download folder'))
            if str(is_contains_new_file) == str(True):
                if  '${is_contains_new_file}'=='0':
                    new_files_name = list_of_current_files[0]
                else:
                    new_files_name = GeneralUtils.check_value_in_file(list_of_existing_files,list_of_current_files)
                new_files_name = BuiltIn().run_keyword_and_return_status(BuiltIn().should_not_match_regexp(new_files_name,'(?i).*\\.tmp','Chrome is still downloading a file'))
            else:
                is_not_temp_file= False
            result = is_contains_new_file and is_not_temp_file
            BuiltIn().exit_for_loop_if(str(result)==str(True))
            BuiltIn().should_be_true(result,'not found file in download_directory')
            file_path = os.path.join(directory,new_files_name)
            return    file_path
    
    # Write new row to excel file
    #write new row at the end to excel file
    #\n ``row_data_as_a_list`` is a list of data represent number of colum in a row ex. ['1','10637','delivered']
    #\n ``excel_file_path`` absolute path to write the data into
    @keyword
    def write_new_row_to_excel_file(self,row_data_as_a_list,excel_file_path):
        date = datetime.datetime.now()
        wb = load_workbook(excel_file_path,date)
        ws = wb.active
        info_from_excel = ws.col(1)
        number_of_row = len(info_from_excel)
        row_number_to_write_to = number_of_row+1
        ws.cell(row=row_number_to_write_to, column=1, value=row_data_as_a_list)
        wb.save(excel_file_path)
        wb.close()

    # Column in excel file should contains correct information
    #[Documentation]     check if column in excel file contains correct information
    #     ...     \n reading column of all rows and expected at least 1 to match expected information
    #     ...     \n ``list_of_expected_information`` is a list contains contain expected information
    #     ...     \n ``column_to_read`` is an integer indicate column number to read
    @keyword
    def column_in_excel_file_should_contains_correct_information(self,excel_file_path,list_of_expected_information,column_to_read):
        date = datetime.datetime.now()
        wb = load_workbook(excel_file_path,date)
        ws = wb.active
        info_from_excel = ws.col(column_to_read)
        Collections.list_should_contain_sub_list(info_from_excel,list_of_expected_information)
        wb.close()

    # Column in exported excel file should exaclty match
    #[Documentation]     check if column in excel file contains correct information
    #     ...     \n reading column of all rows and expected all to match expected information
    #     ...     \n ``list_of_expected_information`` is a list contains contain expected information
    #     ...     \n ``column_to_read`` is an integer indicate column number to read
    @keyword
    def column_in_exported_excel_file_should_exaclty_match(self,excel_file_path,list_of_expected_information,column_to_read):
        date = datetime.datetime.now()
        wb = load_workbook(excel_file_path,date)
        print(wb)
        ws = wb.active
        info_from_excel = ws.col(column_to_read)
        print(info_from_excel)
        Collections.lists_should_be_equal(info_from_excel,list_of_expected_information)
        wb.close()

    #Get body and link from email
    #     [Documentation]     get email body and link from any email providers using imaplibrary.
    #     ...     \n filter using sender address   
    #     ...     \n ``timeout`` how long in second you want to wait for the email
    @keyword
    def get_body_and_link_from_email(self,email,password,sender_address,timeout=60):
        BuiltIn().wait_until_keyword_succeeds('40x','10s')
        ImapLibrary2().open_mailbox('imap.gmail.com',email,password)
        index = ImapLibrary2().wait_for_email(sender_address,'UNSEEN',timeout)
        parts = ImapLibrary2().walk_multipart_email(index)
        for i in range(parts):
            ImapLibrary2().walk_multipart_email(index)
            content_type = ImapLibrary2().get_multipart_content_type
            BuiltIn().continue_for_loop_if(content_type != 'text/html')
            payload = ImapLibrary2().get_multipart_payload(decode=True)
            link = ImapLibrary2().get_links_from_email(index)
        ImapLibrary2().delete_all_emails
        ImapLibrary2().close_mailbox
        links = GeneralUtils.decode_url(link)
        return payload,links
    
    #Get email body
    #[Documentation]     get email body from any email providers using imaplibrary.
    #     ...    \n filter using sender address
    #     ...    \n ``timeout`` how long in second you want to wait for the email
    @keyword
    def get_email_body(self,email,password,sender_address,timeout=60):
        BuiltIn().wait_until_keyword_succeeds('40x','10s')
        ImapLibrary2().open_mailbox('imap.gmail.com',email,password)
        index = ImapLibrary2().wait_for_email(sender_address,'UNSEEN',timeout)
        text = ImapLibrary2().get_email_body(index)
        text_replace = text.replace('\n','')
        ImapLibrary2().delete_all_emails
        ImapLibrary2().close_mailbox
        return text_replace

# if __name__ == "__main__":
#     # webKeyword.lib_click_visible_element('id="CRO-Hero-Button"')
#     print("__main__")
#     # print(f"{keywords}")
#     # print(f"{keywords.Get WebElement()}")